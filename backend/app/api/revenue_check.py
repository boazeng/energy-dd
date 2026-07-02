"""בדיקת הכנסות — קריאת אקסלים מ-SharePoint + השוואות."""
from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import APIRouter, HTTPException
from sqlalchemy import select

from app.core.config import settings
from app.integrations.sharepoint import SharePointError, fetch_file, list_folder

router = APIRouter(prefix="/api/revenue-check", tags=["revenue-check"])

FOLDER_PATH = "כספים/תכנית עסקית/אקסל תוכנית עיסקית/ש.א.ר מוביליטי בעמ/DD/בדיקת הכנסות"
VAT = 1.18  # מע"מ 18%

# ─── מיפויי שמות ─────────────────────────────────────────────────────────────
# קובץ 2 (פירוט לקוח): שם אתר → מילות מפתח בשם הבניין במערכת
SITE_KEYWORDS_FILE2: dict[str, list[str]] = {
    "אייזנברג":                 ["אייזנברג"],
    "אלקבץ":                    ["אלקבץ"],
    "בוסתן 7":                  ["בוסתן"],
    "בלפור 7":                  ["בלפור"],
    "ג'ו עמר 4":                ["ג'ו עמר", "גן עמר"],
    "דודו דותן 3":              ["דודו דותן"],
    "הנחשול 12":                ["נחשול 12"],
    "הרצוג 17 ראש העין":        ["הרצוג"],
    "יהודה הלוי 21+23, יבנה":   ["יהודה הלוי"],
    "כינור 4":                  ["כינור"],
    "נחשול 14":                 ["נחשול 14"],
    'צה"ל 5':                   ['צה"ל', "צהל"],
    "שאולי 17":                 ["שאולי"],
    "שד היובל":                 ["שדרות היובל", "שד היובל", "שד׳ היובל"],
    "שולמית אלוני 3,5":         ["שולמית"],
    "תפוז 13":                  ["תפוז"],
    "היי גרופ":                 ["בן גוריון", "גבע 2", "אשתאול", "קין קאורין"],
}

# קובץ 1 (נתוני בניינים): שם בניין בקובץ → מילות מפתח לשם בניין במערכת
# שמות בקובץ 1 כוללים עיר כקידומת: "אשדוד - שאולי 17", "High Group", וכו'
BUILDING_KEYWORDS_FILE1: dict[str, list[str]] = {
    "אייזנברג":          ["אייזנברג"],
    "אלקבץ":             ["אלקבץ"],
    "בוסתן":             ["בוסתן"],
    "בלפור":             ["בלפור"],
    "ג'ו עמר":           ["ג'ו עמר", "גן עמר"],
    "דודו דותן":         ["דודו דותן"],
    "הנחשול 12":         ["נחשול 12"],
    "הרצוג":             ["הרצוג"],
    "יהודה הלוי":        ["יהודה הלוי"],
    "כינור":             ["כינור"],
    "נחשול 14":          ["נחשול 14"],
    'צה"ל':              ['צה"ל', "צהל"],
    "שאולי":             ["שאולי"],
    "היובל":             ["שדרות היובל", "שד היובל", "שד׳ היובל"],
    "שולמית":            ["שולמית"],
    "תפוז":              ["תפוז"],
    "high group":        ["בן גוריון", "גבע 2", "אשתאול", "קין קאורין"],
}

COMMERCIAL_SITES = {"ארנה נהריה", "סטאר סנטר"}
HI_GROUP_SITE = "היי גרופ"


# ─── עזרים כלליים ────────────────────────────────────────────────────────────

def _parse_excel(content: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append([str(cell) if cell is not None else None for cell in row])
        sheets.append({"sheet": name, "rows": rows})
    return sheets


def _load_excel_files() -> list[dict]:
    try:
        files = list_folder(FOLDER_PATH)
    except SharePointError as e:
        raise HTTPException(status_code=502, detail=str(e))

    result = []
    for f in files:
        if f["is_folder"] or not f["name"].lower().endswith((".xlsx", ".xls")):
            continue
        try:
            content = fetch_file(f["download_url"])
            sheets = _parse_excel(content)
            result.append({"name": f["name"], "web_url": f["web_url"], "sheets": sheets, "error": None})
        except Exception as e:
            result.append({"name": f["name"], "web_url": f.get("web_url", ""), "sheets": [], "error": str(e)})
    return result


def _find_detailed_file(files: list[dict]) -> dict | None:
    """קובץ 2: יש עמודת Driver."""
    for f in files:
        for sheet in f.get("sheets", []):
            rows = sheet.get("rows", [])
            if rows and "Driver" in (rows[0] or []):
                return {"file": f, "sheet": sheet}
    return None


def _find_building_file(files: list[dict]) -> dict | None:
    """קובץ 1: יש עמודת חשמל (Kw) ו-סה"כ פרימיה."""
    for f in files:
        for sheet in f.get("sheets", []):
            rows = sheet.get("rows", [])
            header = rows[0] if rows else []
            if header and any("פרימיה" in (h or "") for h in header):
                return {"file": f, "sheet": sheet}
    return None


def _num(v) -> float:
    try:
        return float(v) if v not in (None, "", "None") else 0.0
    except (ValueError, TypeError):
        return 0.0


def _parse_customers(sheet: dict, month_filter: str | None = None) -> list[dict]:
    rows = sheet["rows"]
    if not rows:
        return []
    header = rows[0]
    idx = {h: i for i, h in enumerate(header) if h}

    def v(row, col):
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    result = []
    for row in rows[1:]:
        date_val = v(row, "Date") or ""
        if month_filter and month_filter not in str(date_val):
            continue
        result.append({
            "id":                   v(row, "Id"),
            "driver":               v(row, "Driver"),
            "site":                 v(row, "Site"),
            "date":                 date_val,
            "monthly_fee_with_vat": _num(v(row, "Monthly fee")),
            "elec_cost":            _num(v(row, "Electricity cost")),
            "due_total":            _num(v(row, "Due total")),
            "status":               v(row, "Status"),
        })
    return result


def _parse_buildings(sheet: dict, month_filter: str | None = None) -> list[dict]:
    """מפרסר שורות מקובץ 1 (נתוני בניינים חודשיים)."""
    rows = sheet["rows"]
    if not rows:
        return []
    header = rows[0]
    prem_col = next((h for h in header if h and "פרימיה" in h), None)
    kwh_col  = next((h for h in header if h and "חשמל" in h and "Kw" in h), None)
    if not prem_col or not kwh_col:
        return []

    idx = {h: i for i, h in enumerate(header) if h}

    def v(row, col):
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    result = []
    for row in rows[1:]:
        date_val = str(v(row, "Date") or "")
        if month_filter and month_filter not in date_val:
            continue
        building = v(row, "בניין/חוזה") or row[2]
        kwh  = _num(v(row, kwh_col))
        prem = _num(v(row, prem_col))
        result.append({
            "building":      building,
            "date":          date_val,
            "kwh":           kwh,
            "premium_incl":  prem,
            "premium_excl":  round(prem / VAT, 2),
            "actual_rate_agorot": round((prem / VAT / kwh * 100) if kwh else 0, 2),
        })
    return result


def _match_bm_names(keywords: list[str], bm_names: list[str]) -> list[str]:
    return [b for b in bm_names if any(kw in b for kw in keywords)]


def _keywords_for_excel1_building(excel_name: str) -> list[str]:
    """מחזיר מילות מפתח לבניין מקובץ 1 לפי substring matching."""
    name_lower = excel_name.lower()
    for key, kws in BUILDING_KEYWORDS_FILE1.items():
        if key.lower() in name_lower:
            return kws
    return []


def _load_projects_charger_counts() -> dict[str, int]:
    path = Path(settings.projects_data_path)
    if not path.is_file():
        return {}
    with path.open(encoding="utf-8") as f:
        data = json.load(f)

    def base(s: str) -> str:
        s = re.sub(r"[0-9]", "", s or "")
        s = re.sub(r'[+\-/,"\']+', " ", s)
        return re.sub(r"\s+", " ", s).strip()

    chargers = data.get("chargers", [])
    buildings = data.get("buildings", [])
    by_base: dict[str, int] = defaultdict(int)
    for c in chargers:
        by_base[base(c.get("project", ""))] += 1

    counts: dict[str, int] = {}
    for b in buildings:
        bname = b.get("project", b.get("name", ""))
        counts[bname] = by_base.get(base(bname), 0)
    return counts


# ─── Endpoints ───────────────────────────────────────────────────────────────

@router.get("/files")
def list_files():
    try:
        files = list_folder(FOLDER_PATH)
    except SharePointError as e:
        raise HTTPException(status_code=502, detail=str(e))
    return {"folder": FOLDER_PATH, "files": files}


@router.get("/data")
def get_excel_data():
    return {"files": _load_excel_files()}


@router.get("/compare/chargers")
def compare_chargers(month: str = "2026-05"):
    """השוואת כמות מטענים: מערכת (projects.json) מול וויבו (נהגים ייחודיים)."""
    files = _load_excel_files()
    found = _find_detailed_file(files)
    if not found:
        raise HTTPException(status_code=404, detail="לא נמצא קובץ פירוט לקוח")

    customers = _parse_customers(found["sheet"], month_filter=month)

    excel_drivers: dict[str, set] = defaultdict(set)
    for c in customers:
        site = (c["site"] or "").strip()
        if site and site not in COMMERCIAL_SITES:
            excel_drivers[site].add(c["driver"])
    excel_counts = {site: len(d) for site, d in excel_drivers.items()}

    sys_counts = _load_projects_charger_counts()
    sys_bm_names = list(sys_counts.keys())

    rows = []
    for excel_site in sorted(excel_counts.keys()):
        excel_n = excel_counts[excel_site]
        kws = SITE_KEYWORDS_FILE2.get(excel_site, [])
        matched = _match_bm_names(kws, sys_bm_names)
        sys_n = sum(sys_counts.get(b, 0) for b in matched) if matched else None
        diff = (excel_n - sys_n) if sys_n is not None else None
        status = (
            "תואם" if diff == 0 else
            "חוסר" if (diff or 0) < 0 else
            "עודף"
        ) if diff is not None else "לא נמצא במערכת"

        rows.append({
            "excel_site":        excel_site,
            "matched_buildings": matched,
            "excel_drivers_may": excel_n,
            "system_chargers":   sys_n,
            "diff":              diff,
            "status":            status,
        })

    return {"month": month, "rows": rows}


@router.get("/compare/monthly-fees")
def compare_monthly_fees(month: str = "2026-05"):
    """השוואת דמי מנוי חודשיים לפי לקוח: וויבו (כולל מע"מ 18%) מול הסכם (ללא מע"מ)."""
    from app.core.db import SessionLocal
    from app.models.building_model import BuildingModel

    with SessionLocal() as db:
        buildings = list(db.scalars(select(BuildingModel)))

    bm_by_name = {b.building_name: b for b in buildings}
    bm_names = list(bm_by_name.keys())

    files = _load_excel_files()
    found = _find_detailed_file(files)
    if not found:
        raise HTTPException(status_code=404, detail="לא נמצא קובץ פירוט לקוח")

    customers = _parse_customers(found["sheet"], month_filter=month)

    rows = []
    for c in customers:
        site = (c["site"] or "").strip()
        if not site or site in COMMERCIAL_SITES:
            continue

        monthly_incl = c["monthly_fee_with_vat"]
        monthly_excl = round(monthly_incl / VAT, 2)

        kws = SITE_KEYWORDS_FILE2.get(site, [])
        matched = _match_bm_names(kws, bm_names)
        expected_fee: float | None = None
        building_display = None

        if matched:
            fees = [bm_by_name[b].mgmt_fee_per_charger for b in matched if b in bm_by_name]
            if fees:
                expected_fee = round(sum(fees) / len(fees), 2)
                building_display = matched[0] if len(matched) == 1 else "היי גרופ"

        diff = round(monthly_excl - expected_fee, 2) if expected_fee is not None else None
        status = (
            "תואם" if diff is not None and abs(diff) < 1 else
            "סטייה" if diff is not None else
            "לא נמצא"
        )

        rows.append({
            "driver":               c["driver"],
            "site":                 site,
            "building":             building_display,
            "date":                 c["date"],
            "monthly_fee_incl_vat": monthly_incl,
            "monthly_fee_excl_vat": monthly_excl,
            "expected_fee":         expected_fee,
            "diff":                 diff,
            "status":               status,
            "pay_status":           c["status"],
        })

    site_summary: dict[str, dict] = {}
    for r in rows:
        s = r["site"]
        if s not in site_summary:
            site_summary[s] = {
                "site": s, "building": r["building"],
                "count": 0, "match": 0, "deviation": 0, "unpaid": 0,
                "expected_fee": r["expected_fee"],
            }
        site_summary[s]["count"] += 1
        if r["status"] == "תואם":   site_summary[s]["match"] += 1
        elif r["status"] == "סטייה": site_summary[s]["deviation"] += 1
        if r["pay_status"] != "Paid": site_summary[s]["unpaid"] += 1

    return {"month": month, "rows": rows, "site_summary": list(site_summary.values())}


@router.get("/compare/kwh-avg")
def compare_kwh_avg():
    """צריכה ממוצעת חודשית לפי בניין ולמטען — ינואר–מאי 2026."""
    from app.core.db import SessionLocal
    from app.models.building_model import BuildingModel

    with SessionLocal() as db:
        buildings = list(db.scalars(select(BuildingModel)))

    bm_by_name = {b.building_name: b for b in buildings}
    bm_names   = list(bm_by_name.keys())

    files = _load_excel_files()
    found = _find_building_file(files)
    if not found:
        raise HTTPException(status_code=404, detail="לא נמצא קובץ נתוני בניינים")

    all_rows = _parse_buildings(found["sheet"], month_filter=None)

    by_bldg: dict[str, list[dict]] = defaultdict(list)
    for br in all_rows:
        bldg = (br["building"] or "").strip()
        if not bldg or any(c in bldg for c in COMMERCIAL_SITES):
            continue
        by_bldg[bldg].append(br)

    result = []
    for excel_name, rows_list in sorted(by_bldg.items()):
        kws     = _keywords_for_excel1_building(excel_name)
        matched = _match_bm_names(kws, bm_names)
        charger_count: int | None = None
        building_display = None
        if matched:
            charger_count    = sum(getattr(bm_by_name.get(b), "current_chargers", 0) for b in matched)
            building_display = matched[0] if len(matched) == 1 else "היי גרופ"

        months_data = []
        for br in sorted(rows_list, key=lambda r: r["date"]):
            kwh     = br["kwh"]
            kwh_per = round(kwh / charger_count, 2) if (charger_count and kwh) else None
            months_data.append({
                "month":           br["date"][:7],
                "kwh":             kwh,
                "kwh_per_charger": kwh_per,
            })

        valid_kwh = [m["kwh"] for m in months_data if m["kwh"]]
        avg_kwh   = round(sum(valid_kwh) / len(valid_kwh), 2) if valid_kwh else None
        avg_per   = round(avg_kwh / charger_count, 2) if (avg_kwh and charger_count) else None

        result.append({
            "building_excel":      excel_name,
            "building_system":     building_display,
            "current_chargers":    charger_count,
            "months":              months_data,
            "avg_monthly_kwh":     avg_kwh,
            "avg_kwh_per_charger": avg_per,
        })

    return {"buildings": result}


@router.get("/compare/electricity")
def compare_electricity(month: str = "2026-05"):
    """השוואת תעריף חשמל (עמלת וויבו) לפי בניין: פרימיה÷מע"מ÷קוט"ש מול הסכם (ללא מע"מ)."""
    from app.core.db import SessionLocal
    from app.models.building_model import BuildingModel

    with SessionLocal() as db:
        buildings = list(db.scalars(select(BuildingModel)))

    bm_by_name = {b.building_name: b for b in buildings}
    bm_names = list(bm_by_name.keys())

    files = _load_excel_files()
    found = _find_building_file(files)
    if not found:
        raise HTTPException(status_code=404, detail="לא נמצא קובץ נתוני בניינים")

    building_rows = _parse_buildings(found["sheet"], month_filter=month)

    rows = []
    for br in building_rows:
        bldg = (br["building"] or "").strip()
        if not bldg or any(c in bldg for c in COMMERCIAL_SITES):
            continue

        kwh    = br["kwh"]
        actual = br["actual_rate_agorot"]

        kws     = _keywords_for_excel1_building(bldg)
        matched = _match_bm_names(kws, bm_names)
        expected_rate: float | None = None
        building_display = None

        if matched:
            rates = [bm_by_name[b].electricity_rate_agorot for b in matched if b in bm_by_name]
            if rates:
                expected_rate    = round(sum(rates) / len(rates), 2)
                building_display = matched[0] if len(matched) == 1 else "היי גרופ"

        diff = round(actual - expected_rate, 2) if expected_rate is not None and actual else None
        status = (
            "תואם"  if diff is not None and abs(diff) < 1 else
            "סטייה" if diff is not None else
            "לא נמצא"
        )

        rows.append({
            "building_excel":     bldg,
            "building_system":    building_display,
            "date":               br["date"],
            "kwh":                kwh,
            "premium_incl_vat":   br["premium_incl"],
            "premium_excl_vat":   br["premium_excl"],
            "actual_rate_agorot": actual,
            "expected_rate_agorot": expected_rate,
            "diff":               diff,
            "status":             status,
        })

    # סיכום לפי בניין (כל החודשים)
    bldg_summary: dict[str, dict] = {}
    for r in rows:
        k = r["building_excel"]
        if k not in bldg_summary:
            bldg_summary[k] = {
                "building_excel":   k,
                "building_system":  r["building_system"],
                "expected_rate":    r["expected_rate_agorot"],
                "months": [], "deviations": 0,
            }
        bldg_summary[k]["months"].append({
            "date": r["date"][:7], "kwh": r["kwh"],
            "actual_rate": r["actual_rate_agorot"], "diff": r["diff"],
        })
        if r["status"] == "סטייה":
            bldg_summary[k]["deviations"] += 1

    return {
        "month": month,
        "rows": rows,
        "building_summary": list(bldg_summary.values()),
    }
